from os import sys, startfile
import ctypes
import psycopg2
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter as letter
from openpyxl.styles import Alignment, Font
from datetime import datetime
from PyQt6.QtCore import QSize, Qt, QRect, QPoint
from PyQt6.QtWidgets import QApplication, QLabel, QWidget, QMainWindow, QPushButton, QVBoxLayout, \
							QFileDialog, QLineEdit, QDateEdit, QComboBox, QTextEdit, QDialog, \
							QMessageBox, QMenu, QCheckBox
from PyQt6.QtGui import QPainter, QColor, QBrush, QFont, QAction

def sort_func(item):
	return item[3], item[4], item[5], datetime.now().date() - item[2]

class MainWindow(QMainWindow):
	def __init__(self):
		super().__init__()

		self.font_arial_size7 = QFont("Arial", 7)
		self.font_arial_size8 = QFont("Arial", 8)
		self.font_arial_size8_bold = QFont("Arial", 8)
		self.font_arial_size8_bold.setBold(True)
		self.font_arial_size10 = QFont("Arial", 10)
		self.font_arial_size12 = QFont("Arial", 12)
		self.font_arial_size15 = QFont("Arial", 15)

		self.setMouseTracking(True)
		self.mouse_pos = QPoint(0, 0)

		self.all_data_saved = []
		self.all_wind_in_body = []
		self.menu_widgets = []
		self.menu_labels = []
		self.report_info_lines = []
		self.report_info_labels = []
		self.page_in_body = 1
		self.free_line_in_body = 0
		self.count_of_lines_in_body = 18
		self.line_changing = 0
		self.id_wind_changing = 0
		self.new_data_for_changing = ''
		self.delete_check = 0
		self.add_check = 2
		self.selection_line = -1

		self.mode_menu = 'searching'
		self.dialog_menu_for_enter_add_button_answer = 1

		self.russian_names_coloms = ["№ протокола", \
									"Дата выдачи", \
									"Фамилия", \
									"Имя", \
									"Отчество", \
									"Дата рождения", \
									"Адрес", \
									"Инвалидность", \
									"ОМС", \
									"№ карты", \
									"№ участка", \
									"Диагноз", \
									"Код", \
									"Фамилия врача", \
									"Тип ф.70", \
									]
		self.english_names_coloms = ["protocol_id", \
									"date_of_issue", \
									"second_name", \
									"first_name", \
									"surname", \
									"date_of_birth", \
									"address", \
									"disability", \
									"CHI", \
									"outpatient_card_id", \
									"site", \
									"diagnosis", \
									"code", \
									"doctor_second_name", \
									"type_f70", \
									]

		self.disability_items = ['-', 'инв I гр.', 'инв II гр.', 'инв III гр.', 'раб.', 'не раб.', 'реб. инв.']
		self.type_f70_items = ['-', 'органы дыхания', 'органы опор-дв. апп.', 'органы ССС', 'органы эндок. сист.', \
								'органы нерв. сист.', 'другое']
		self.type_report_items = ['Взрослые', 'Дети', 'Инвалиды']
		self.year_report_items = [str(year) for year in range(int(datetime.now().strftime("%Y")), 1999, -1)]
		self.period_report_items = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', \
									'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь', 'Январь - Март', \
									'Январь - Июнь', 'Январь - Сентябрь', 'Год']

		self.report_type = self.type_report_items[0]
		self.report_year = self.year_report_items[0]
		self.report_period = 1
		self.report_period_type = 'mons'
		self.report_type_f70 = ''
		self.report_info = [0, 0, 0, 0, 0, 0, 0]


		self.x_len_labels = []
		for i in range(15):
			if i == 4: self.x_len_labels += [120]
			elif i == 6: self.x_len_labels += [150]
			elif i == 8: self.x_len_labels += [125]
			elif i == 9: self.x_len_labels += [80]
			elif i == 11: self.x_len_labels += [130]
			elif i == 12: self.x_len_labels += [60]
			elif i == 14: self.x_len_labels += [130]
			else: self.x_len_labels += [100]

		self.setWindowTitle("Database")
		self.setFixedSize(QSize(1700, 800))
		self.setMinimumSize(800, 600)

		user32 = ctypes.windll.user32
		self.screen_size = (user32.GetSystemMetrics(0), user32.GetSystemMetrics(1))

		self.create_search_button()
		self.create_add_button()
		self.create_report_button()
		self.create_labels()
		self.create_all_wind_in_body()


		self.protocol_id_data = ''
		self.date_of_issue_data = ''
		self.first_name_data = ''
		self.second_name_data = ''
		self.surname_data = ''
		self.date_of_birth_data = ''
		self.address_data = ''
		self.disability_data = ''
		self.CHI_data = ''
		self.outpatient_card_id_data = ''
		self.site_data = ''
		self.diagnosis_data = ''
		self.code_data = ''
		self.doctor_second_name_data = ''
		self.type_f70_data = ''

		self.menu()

	def create_labels(self):
		x_coord = 20
		for i in range(15):
			body_label = QLabel(self.russian_names_coloms[i], self)
			body_label.setGeometry(QRect(x_coord, 225, self.x_len_labels[i], 20))
			body_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
			body_label.setFont(self.font_arial_size10)
			body_label.setStyleSheet("background-color:" + "#eeeeee; border: 1px solid black;")

			x_coord += self.x_len_labels[i] + 5

	def create_search_button(self):
		self.search_button = QPushButton(self)
		self.search_button.setCheckable(True)
		self.search_button.setGeometry(QRect(0, 0, 50, 20))
		self.search_button.setText("Поиск")
		self.search_button.clicked.connect(self.the_search_button_was_clicked)

	def create_add_button(self):
		self.add_button = QPushButton(self)
		self.add_button.setCheckable(True)
		self.add_button.setGeometry(QRect(50, 0, 65, 20))
		self.add_button.setText("Добавить")
		self.add_button.clicked.connect(self.the_add_button_was_clicked)

	def create_report_button(self):
		self.report_button = QPushButton(self)
		self.report_button.setCheckable(True)
		self.report_button.setGeometry(QRect(115, 0, 50, 20))
		self.report_button.setText("Отчеты")
		self.report_button.clicked.connect(self.the_report_button_was_clicked)

	def create_all_wind_in_body(self):

		for j in range(self.count_of_lines_in_body):
			x_coord = 20
			one_line = []

			for i in range(15):
				
				print_data_line = QTextEdit(self)

				print_data_line.setGeometry(QRect(x_coord, 255 + 30*j, self.x_len_labels[i], 25))
				print_data_line.setAlignment(Qt.AlignmentFlag.AlignCenter)
				print_data_line.setMouseTracking(True)
				print_data_line.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
				print_data_line.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
				print_data_line.customContextMenuRequested.connect(self.custom_context_menu_for_body_data)

				if i in [2, 3, 4, 6, 8, 11]: print_data_line.setFont(self.font_arial_size8)
				else: print_data_line.setFont(self.font_arial_size10)

				print_data_line.setStyleSheet("border: 1px solid black; background-color: #888888;")
				print_data_line.setReadOnly(True)
				print_data_line.hide()

				one_line += [print_data_line]
				x_coord += self.x_len_labels[i] + 5

			self.all_wind_in_body += [one_line]

	def input_error_wind(self, err):

		self.dialog_error = QDialog(self)
		self.dialog_error.setGeometry(QRect(self.screen_size[0]//2 - 115, self.screen_size[1]//2 - 50, 230, 100))
		self.dialog_error.setWindowTitle("Ошибка ввода")

		label_1 = QLabel("Неверно введено поле:", self.dialog_error)
		label_1.setAlignment(Qt.AlignmentFlag.AlignCenter)
		label_1.setFont(self.font_arial_size10)
		label_1.setGeometry(QRect(0, 10, 230, 20))

		label_2 = QLabel(self.russian_names_coloms[err], self.dialog_error)
		label_2.setAlignment(Qt.AlignmentFlag.AlignCenter)
		label_2.setFont(self.font_arial_size12)
		label_2.setGeometry(QRect(0, 35, 230, 20))

		ok_button = QPushButton(self.dialog_error)
		ok_button.setCheckable(True)
		ok_button.setGeometry(QRect(75, self.dialog_error.size().height() - 30, 80, 25))
		ok_button.setText("OK")
		ok_button.clicked.connect(lambda: self.dialog_error.hide())

		self.dialog_error.exec()

	def cheking_correctness_of_data(self, data):

		result = -1
		for index in range(15):
			if result != -1: break

			if index == 0:
				if len(data[index]) > 10: result = index
				try:
					data_split = data[index].split("/")
					if len(data_split) != 2: result = index
					data_1 = int(data_split[0])
					data_1 = int(data_split[1])
				except ValueError: result = index

			elif index in [2, 3, 4]:
				if data[index] == '': result = index
				if len(data[index]) > 20: result = index

			elif index in [1, 5, 7, 14] and data[index] == '': result = index
			elif index == 6 and len(data[index]) > 100: result = index

			elif index == 8:
				if data[index] == '': continue
				if len(data[index]) != 16: result = index
				try:
					for i in range(16):
						int(data[index][i])
				except ValueError: result = index

			elif index == 9:
				if data[index] == '': continue
				if len(data[index]) > 4: result = index
				try:
					for i in range(len(data[index])):
						int(data[index][i])
				except ValueError: result = index

			elif index == 10:
				if data[index] == '': continue
				if len(data[index]) > 3: result = index
				try:
					for i in range(len(data[index])):
						int(data[index][i])
				except ValueError: result = index

			elif index == 11 and (data[index] == '' or len(data[index]) > 50): result = index

			elif index == 12:
				if data[index] == '': result = index
				if len(data[index]) > 10: result = index

			elif index == 13 and len(data[index]) > 20: result = index

		if result != -1:
			self.input_error_wind(result)
			return False
		else: return True

	def menu(self):

		x_coord = 20
		for i in range(15):

			if self.english_names_coloms[i] == 'disability':
				input_line = QComboBox(self)
				input_line.addItems(self.disability_items)
				self.menu_widgets += [input_line]
				input_line.activated.connect(getattr(self, 'get_' + self.english_names_coloms[i] + '_data'))
				input_line.setFont(self.font_arial_size10)

			elif self.english_names_coloms[i] == 'type_f70':
				input_line = QComboBox(self)
				input_line.addItems(self.type_f70_items)
				self.menu_widgets += [input_line]
				input_line.activated.connect(getattr(self, 'get_' + self.english_names_coloms[i] + '_data'))
				input_line.setFont(self.font_arial_size10)

			elif self.english_names_coloms[i] == 'second_name' or self.english_names_coloms[i] == 'first_name' \
				or self.english_names_coloms[i] == 'surname' or self.english_names_coloms[i] == 'address' \
				or self.english_names_coloms[i] == 'diagnosis':

				input_line = QTextEdit(self)
				self.menu_widgets += [input_line]
				input_line.textChanged.connect(getattr(self, 'get_' + self.english_names_coloms[i] + '_data'))
				input_line.setAlignment(Qt.AlignmentFlag.AlignCenter)
				input_line.setFont(self.font_arial_size8)
				input_line.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

			else:
				input_line = QLineEdit(self)
				self.menu_widgets += [input_line]
				input_line.textEdited.connect(getattr(self, 'get_' + self.english_names_coloms[i] + '_data'))
				input_line.setAlignment(Qt.AlignmentFlag.AlignCenter)
				input_line.setFont(self.font_arial_size10)


			input_line.setGeometry(QRect(x_coord, 30, self.x_len_labels[i], 25))
			input_line.setStyleSheet("border: 2px solid black;")

			input_label = QLabel(self.russian_names_coloms[i], self)
			input_label.setGeometry(QRect(x_coord, 50, self.x_len_labels[i], 30))
			input_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
			input_label.setFont(self.font_arial_size8)
			self.menu_labels += [input_label]

			x_coord += self.x_len_labels[i] + 5


		self.report_type_line = QComboBox(self)
		self.report_type_line.addItems(self.type_report_items)
		self.report_type_line.activated.connect(self.get_report_type_data)
		self.report_type_line.setFont(self.font_arial_size10)
		self.report_type_line.setGeometry(QRect(20, 30, 100, 25))
		self.report_type_line.setStyleSheet("background-color: #57cf77; border: 2px solid black;")
		self.report_type_line.hide()

		self.report_year_line = QComboBox(self)
		self.report_year_line.addItems(self.year_report_items)
		self.report_year_line.activated.connect(self.get_report_year_data)
		self.report_year_line.setFont(self.font_arial_size10)
		self.report_year_line.setGeometry(QRect(130, 30, 100, 25))
		self.report_year_line.setStyleSheet("background-color: #57cf77; border: 2px solid black;")
		self.report_year_line.hide()

		self.report_period_line = QComboBox(self)
		self.report_period_line.addItems(self.period_report_items)
		self.report_period_line.activated.connect(self.get_report_period_data)
		self.report_period_line.setFont(self.font_arial_size10)
		self.report_period_line.setGeometry(QRect(240, 30, 140, 25))
		self.report_period_line.setStyleSheet("background-color: #57cf77; border: 2px solid black;")
		self.report_period_line.hide()

		self.report_type_f70_line = QComboBox(self)
		self.report_type_f70_line.addItems(self.type_f70_items)
		self.report_type_f70_line.activated.connect(self.get_report_type_f70_data)
		self.report_type_f70_line.setFont(self.font_arial_size10)
		self.report_type_f70_line.setGeometry(QRect(390, 30, 130, 25))
		self.report_type_f70_line.setStyleSheet("border: 2px solid black;")
		self.report_type_f70_line.hide()


		for i in range(7):	
			info_line = QLineEdit(self)
			info_line.setAlignment(Qt.AlignmentFlag.AlignCenter)
			info_line.setFont(self.font_arial_size10)
			info_line.setGeometry(QRect(20 + 120*i, 65, 110, 25))
			if i == 0: info_line.setStyleSheet("border: 2px solid black;")
			else: info_line.setStyleSheet("border: 1px solid black;")
			info_line.setReadOnly(True)
			info_line.hide()

			if i == 0: 
				info_label = QLabel('Всего', self)
				info_label.setFont(self.font_arial_size8_bold)
			else: 
				info_label = QLabel(self.type_f70_items[i], self)
				info_label.setFont(self.font_arial_size8)
			info_label.setGeometry(QRect(20 + 120*i, 90, 110, 30))
			info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
			info_label.hide()

			self.report_info_lines += [info_line]
			self.report_info_labels += [info_label]


		self.enter_search_button = QPushButton(self)
		self.enter_search_button.setCheckable(True)
		self.enter_search_button.setGeometry(QRect(self.size().width() - 145, 170, 130, 35))
		self.enter_search_button.setText("Искать")
		self.enter_search_button.clicked.connect(self.the_enter_search_button_was_clicked)
		self.enter_search_button.setFont(self.font_arial_size15)
		self.enter_search_button.setMouseTracking(True)
		self.enter_search_button.setStyleSheet("background-color: #009900; border-radius: 10px; \
												border: 1px solid black;")

		self.enter_add_button = QPushButton(self)
		self.enter_add_button.setCheckable(True)
		self.enter_add_button.setGeometry(QRect(self.size().width() - 145, 170, 130, 35))
		self.enter_add_button.setText("Добавить")
		self.enter_add_button.clicked.connect(self.the_enter_add_button_was_clicked)
		self.enter_add_button.setFont(self.font_arial_size15)
		self.enter_add_button.setMouseTracking(True)
		self.enter_add_button.setStyleSheet("background-color: #0000aa; border-radius: 10px; \
												border: 1px solid black;")
		self.enter_add_button.hide()

		self.enter_report_button = QPushButton(self)
		self.enter_report_button.setCheckable(True)
		self.enter_report_button.setGeometry(QRect(self.size().width() - 145, 170, 130, 35))
		self.enter_report_button.setText("Расшифровать")
		self.enter_report_button.clicked.connect(self.the_enter_report_button_was_clicked)
		self.enter_report_button.setFont(self.font_arial_size12)
		self.enter_report_button.setMouseTracking(True)
		self.enter_report_button.setStyleSheet("background-color: #f7ca4d; border-radius: 10px; \
												border: 1px solid black;")
		self.enter_report_button.hide()

		self.report_in_excel_button = QPushButton(self)
		self.report_in_excel_button.setCheckable(True)
		self.report_in_excel_button.setGeometry(QRect(20, 170, 100, 35))
		self.report_in_excel_button.setText("Отчет в Excel")
		self.report_in_excel_button.clicked.connect(self.the_report_in_excel_button_was_clicked)
		self.report_in_excel_button.setFont(self.font_arial_size10)
		self.report_in_excel_button.setMouseTracking(True)
		self.report_in_excel_button.setStyleSheet("background-color: #bbbbbb; border-radius: 10px; \
												border: 1px solid black;")
		self.report_in_excel_button.hide()
		
		self.clear_all_menu_button = QPushButton(self)
		self.clear_all_menu_button.setCheckable(True)
		self.clear_all_menu_button.setGeometry(QRect(20, 170, 100, 35))
		self.clear_all_menu_button.setText("Очистить все поля")
		self.clear_all_menu_button.clicked.connect(self.the_clear_all_menu_button_was_clicked)
		self.clear_all_menu_button.setFont(self.font_arial_size7)
		self.clear_all_menu_button.setMouseTracking(True)
		self.clear_all_menu_button.setStyleSheet("background-color: #bbbbbb; border-radius: 10px; \
												border: 1px solid black;")

		self.show_all_data_in_db_button = QPushButton(self)
		self.show_all_data_in_db_button.setCheckable(True)
		self.show_all_data_in_db_button.setGeometry(QRect(20, 130, 100, 35))
		self.show_all_data_in_db_button.setText("Показать всю таблицу")
		self.show_all_data_in_db_button.clicked.connect(self.the_show_all_data_in_db_button_was_clicked)
		self.show_all_data_in_db_button.setFont(self.font_arial_size7)
		self.show_all_data_in_db_button.setMouseTracking(True)
		self.show_all_data_in_db_button.setStyleSheet("background-color: #bbbbbb; border-radius: 10px; \
												border: 1px solid black;")

		self.page_down_button = QPushButton(self)
		self.page_down_button.setCheckable(True)
		self.page_down_button.setGeometry(QRect(self.size().width() - 260, 170, 35, 35))
		self.page_down_button.setText("▼")
		self.page_down_button.clicked.connect(self.the_page_down_button_was_clicked)
		self.page_down_button.setFont(self.font_arial_size10)
		self.page_down_button.setMouseTracking(True)
		self.page_down_button.setStyleSheet("background-color: #bbbbbb; border-radius: 10px; \
												border: 1px solid black;")

		self.page_up_button = QPushButton(self)
		self.page_up_button.setCheckable(True)
		self.page_up_button.setGeometry(QRect(self.size().width() - 300, 170, 35, 35))
		self.page_up_button.setText("▲")
		self.page_up_button.clicked.connect(self.the_page_up_button_was_clicked)
		self.page_up_button.setFont(self.font_arial_size10)
		self.page_up_button.setMouseTracking(True)
		self.page_up_button.setStyleSheet("background-color: #bbbbbb; border-radius: 10px; \
												border: 1px solid black;")

		self.page_line = QLineEdit(self)
		self.page_line.setAlignment(Qt.AlignmentFlag.AlignRight)
		self.page_line.setFont(self.font_arial_size12)
		self.page_line.setGeometry(QRect(self.size().width() - 210, 170, 20, 35))
		self.page_line.setStyleSheet("background-color: #969696; border: 0px solid black;")
		self.page_line.setText(str(self.page_in_body))
		self.page_line.textEdited.connect(self.change_page)

		self.page_label = QLabel("/ " + str(len(self.all_data_saved) // (self.count_of_lines_in_body+1) + 1), self)
		self.page_label.setGeometry(QRect(self.size().width() - 190, 170, 20, 35))
		self.page_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
		self.page_label.setFont(self.font_arial_size12)

	def mouseMoveEvent(self, e):

		self.mouse_pos = e.position().toPoint()

		if self.mouse_pos in QRect(20, 170, 100, 35):
			self.clear_all_menu_button.setStyleSheet("background-color: #777777; border-radius: 10px; \
														border: 1px solid black;")
			self.report_in_excel_button.setStyleSheet("background-color: #777777; border-radius: 10px; \
														border: 1px solid black;")
		elif self.mouse_pos not in QRect(20, 170, 100, 35): 
			self.clear_all_menu_button.setStyleSheet("background-color: #bbbbbb; border-radius: 10px; \
														border: 1px solid black;")
			self.report_in_excel_button.setStyleSheet("background-color: #bbbbbb; border-radius: 10px; \
														border: 1px solid black;")

		if self.mouse_pos in QRect(20, 130, 100, 35):
			self.show_all_data_in_db_button.setStyleSheet("background-color: #777777; border-radius: 10px; \
														border: 1px solid black;")
		elif self.mouse_pos not in QRect(20, 130, 100, 35): 
			self.show_all_data_in_db_button.setStyleSheet("background-color: #bbbbbb; border-radius: 10px; \
														border: 1px solid black;")
		
		if self.mouse_pos in QRect(self.size().width() - 260, 170, 35, 35):
			self.page_down_button.setStyleSheet("background-color: #777777; border-radius: 10px; \
														border: 1px solid black;")
		elif self.mouse_pos not in QRect(self.size().width() - 260, 170, 35, 35): 
			self.page_down_button.setStyleSheet("background-color: #bbbbbb; border-radius: 10px; \
														border: 1px solid black;")

		if self.mouse_pos in QRect(self.size().width() - 300, 170, 35, 35):
			self.page_up_button.setStyleSheet("background-color: #777777; border-radius: 10px; \
														border: 1px solid black;")
		elif self.mouse_pos not in QRect(self.size().width() - 300, 170, 35, 35): 
			self.page_up_button.setStyleSheet("background-color: #bbbbbb; border-radius: 10px; \
														border: 1px solid black;")


		if self.mode_menu == 'searching':
			if self.mouse_pos in QRect(self.size().width() - 145, 170, 130, 35):
				self.enter_search_button.setStyleSheet("background-color: #007700; border-radius: 10px; \
														border: 1px solid black;")
			elif self.mouse_pos not in QRect(self.size().width() - 145, 170, 130, 35): 
				self.enter_search_button.setStyleSheet("background-color: #009900; border-radius: 10px; \
														border: 1px solid black;")
		elif self.mode_menu == 'adding':
			if self.mouse_pos in QRect(self.size().width() - 145, 170, 130, 35):
				self.enter_add_button.setStyleSheet("background-color: #000099; border-radius: 10px; \
														border: 1px solid black;")
			elif self.mouse_pos not in QRect(self.size().width() - 145, 170, 130, 35): 
				self.enter_add_button.setStyleSheet("background-color: #0000aa; border-radius: 10px; \
														border: 1px solid black;")

		elif self.mode_menu == 'reporting':
			if self.mouse_pos in QRect(self.size().width() - 145, 170, 130, 35):
				self.enter_report_button.setStyleSheet("background-color: #755705; border-radius: 10px; \
														border: 1px solid black;")
			elif self.mouse_pos not in QRect(self.size().width() - 145, 170, 130, 35): 
				self.enter_report_button.setStyleSheet("background-color: #f7ca4d; border-radius: 10px; \
														border: 1px solid black;")

	def keyPressEvent(self, e):

		if e.modifiers() == Qt.KeyboardModifier.ControlModifier:
			if e.key() == Qt.Key.Key_Enter.value - 1:
				if self.mode_menu == 'searching': self.the_enter_search_button_was_clicked()
				elif self.mode_menu == 'adding': self.the_enter_add_button_was_clicked()
				elif self.mode_menu == 'reporting': self.the_enter_report_button_was_clicked()
			elif e.key() == Qt.Key.Key_Space.value: self.the_clear_all_menu_button_was_clicked()
		elif e.key() == Qt.Key.Key_Down.value: self.the_page_down_button_was_clicked()
		elif e.key() == Qt.Key.Key_Up.value:  self.the_page_up_button_was_clicked()

	'''
	def mousePressEvent(self, e):
		print(e.position())
		print(e.button() == Qt.MouseButton.RightButton)
	'''

	def get_report_type_data(self, data):
		self.report_type = self.type_report_items[data]
		for i in range(7):
			self.report_info_lines[i].setText('')

	def get_report_year_data(self, data):
		self.report_year = self.year_report_items[data]
		for i in range(7):
			self.report_info_lines[i].setText('')

	def get_report_period_data(self, data):
		if data <= 11: 
			self.report_period = data + 1
			self.report_period_type = 'mons'
		else:
			self.report_period_type = 'period'
			if data == 12: self.report_period = 3
			elif data == 13: self.report_period = 6
			elif data == 14: self.report_period = 9
			elif data == 15: self.report_period = 12

		for i in range(7):
			self.report_info_lines[i].setText('')

	def get_report_type_f70_data(self, data):
		if data == 0: self.report_type_f70 = ''
		else: self.report_type_f70 = self.type_f70_items[data]

	def custom_context_menu_for_body_data(self, pos):

		sender = self.sender()

		for line in range(self.count_of_lines_in_body):
			if sender in self.all_wind_in_body[line]:
				self.line_changing = line
				self.id_wind_changing = self.all_wind_in_body[line].index(sender)
				break

		self.selection_line = self.line_changing

		self.dialog = QDialog(self)
		self.dialog.setGeometry(QRect(self.screen_size[0]//2 - 155, self.screen_size[1]//2 - 100, 310, 200))
		self.dialog.setWindowTitle("Изменение")

		label_1 = QLabel("Старые данные: ", self.dialog)
		label_1.setAlignment(Qt.AlignmentFlag.AlignCenter)
		label_1.setFont(self.font_arial_size10)
		label_1.setGeometry(QRect(0, 15, 120, 28))

		line_1 = QTextEdit(sender.toPlainText(), self.dialog)
		line_1.setGeometry(QRect(120, 15, 180, 28))
		line_1.setAlignment(Qt.AlignmentFlag.AlignCenter)
		line_1.setFont(self.font_arial_size12)
		line_1.setStyleSheet("background-color: #ffffff; border: 1px solid black;")
		line_1.setReadOnly(True)
		line_1.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

		label_2 = QLabel("Новые данные: ", self.dialog)
		label_2.setAlignment(Qt.AlignmentFlag.AlignCenter)
		label_2.setFont(self.font_arial_size10)
		label_2.setGeometry(QRect(0, 50, 120, 28))

		if self.id_wind_changing == 7:
			self.line_for_changing = QComboBox(self.dialog)
			self.line_for_changing.addItems(self.disability_items)
			self.line_for_changing.activated.connect(self.new_data_for_changing_in_index)

		elif self.id_wind_changing == 14:
			self.line_for_changing = QComboBox(self.dialog)
			self.line_for_changing.addItems(self.type_f70_items)
			self.line_for_changing.activated.connect(self.new_data_for_changing_in_index)

		else:
			self.line_for_changing = QTextEdit(self.dialog)
			self.line_for_changing.setAlignment(Qt.AlignmentFlag.AlignCenter)
			self.line_for_changing.textChanged.connect(self.new_data_for_changing_in_string_or_date)
			self.line_for_changing.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

		self.line_for_changing.setGeometry(QRect(120, 50, 180, 28))
		self.line_for_changing.setFont(self.font_arial_size12)
		self.line_for_changing.setStyleSheet("background-color: #ffffff; border: 1px solid black;")

		label_3 = QLabel("Удалить старые данные", self.dialog)
		label_3.setAlignment(Qt.AlignmentFlag.AlignCenter)
		label_3.setFont(self.font_arial_size10)
		label_3.setGeometry(QRect(0, 85, 160, 28))

		check_box_1 = QCheckBox(self.dialog)
		check_box_1.setGeometry(QRect(165, 87, 100, 28))
		check_box_1.stateChanged.connect(self.delete_check_data)

		label_4 = QLabel("Добавить новые данные", self.dialog)
		label_4.setAlignment(Qt.AlignmentFlag.AlignCenter)
		label_4.setFont(self.font_arial_size10)
		label_4.setGeometry(QRect(0, 115, 160, 28))

		check_box_2 = QCheckBox(self.dialog)
		check_box_2.setGeometry(QRect(165, 117, 100, 28))
		check_box_2.setChecked(True)
		check_box_2.stateChanged.connect(self.add_check_data)

		accept_button = QPushButton(self.dialog)
		accept_button.setCheckable(True)
		accept_button.setGeometry(QRect(30, self.dialog.size().height() - 40, 110, 30))
		accept_button.setText("Принять")
		accept_button.setFont(self.font_arial_size10)
		accept_button.clicked.connect(self.changing_data)

		cancel_button = QPushButton(self.dialog)
		cancel_button.setCheckable(True)
		cancel_button.setGeometry(QRect(self.dialog.size().width() - 140, self.dialog.size().height() - 40, 110, 30))
		cancel_button.setText("Отмена")
		cancel_button.setFont(self.font_arial_size10)
		cancel_button.clicked.connect(lambda: self.dialog.hide())

		self.dialog.exec()

		self.new_data_for_changing = ''
		self.delete_check = 0
		self.add_check = 2
		self.selection_line = -1

	def new_data_for_changing_in_string_or_date(self):

		if self.id_wind_changing in [1, 5]:
			try: self.new_data_for_changing = datetime.strptime(self.line_for_changing.toPlainText(), '%d.%m.%Y').date()
			except ValueError: self.new_data_for_changing = ''
		elif self.id_wind_changing in [2, 3, 4, 13]: self.new_data_for_changing = self.line_for_changing.toPlainText().title()
		elif self.id_wind_changing == 6: self.new_data_for_changing = self.line_for_changing.toPlainText().capitalize()
		elif self.id_wind_changing == 11: self.new_data_for_changing = self.line_for_changing.toPlainText().lower()
		else: self.new_data_for_changing = self.line_for_changing.toPlainText()

	def new_data_for_changing_in_index(self, data):
		if self.id_wind_changing == 7:
			if data == 0: self.new_data_for_changing = ''
			else: self.new_data_for_changing = self.disability_items[data]
		elif self.id_wind_changing == 14: 
			if data == 0: self.new_data_for_changing = ''
			else: self.new_data_for_changing = self.type_f70_items[data]

	def delete_check_data(self, data):
		self.delete_check = data

	def add_check_data(self, data):
		self.add_check = data

	def changing_data(self):

		index_of_line = self.line_changing + (self.page_in_body - 1) * self.count_of_lines_in_body

		if self.delete_check == 2 and self.add_check == 0:
			sql_request = "DELETE FROM f70 WHERE id = %s"
			cursor.execute(sql_request, [self.all_data_saved[index_of_line][0]])
			self.all_data_saved.remove(self.all_data_saved[index_of_line])

			self.page_label.setText("/ " + str(len(self.all_data_saved) // (self.count_of_lines_in_body+1) + 1))
			if self.page_in_body > len(self.all_data_saved) // (self.count_of_lines_in_body + 1) + 1: self.page_in_body -= 1
			self.page_line.setText(str(self.page_in_body))

		elif self.delete_check == 2 and self.add_check == 2:

			new_line = []
			for i in range(16):
				if i == self.id_wind_changing + 1: new_line += [self.new_data_for_changing]
				else: new_line += [self.all_data_saved[index_of_line][i]]

			if self.cheking_correctness_of_data(new_line[1:]) == False: return

			sql_request = "UPDATE f70 \
							SET " + self.english_names_coloms[self.id_wind_changing] + " = %s \
							WHERE id = %s "

			cursor.execute(sql_request, [self.new_data_for_changing, self.all_data_saved[index_of_line][0]])

			if self.id_wind_changing in [1, 5]:
				self.all_wind_in_body[self.line_changing][self.id_wind_changing].setText(\
					self.new_data_for_changing.strftime("%d.%m.%Y"))
			else:
				self.all_wind_in_body[self.line_changing][self.id_wind_changing].setText(self.new_data_for_changing)
			self.all_wind_in_body[self.line_changing][self.id_wind_changing].setAlignment(Qt.AlignmentFlag.AlignCenter)

			self.all_data_saved[index_of_line] = tuple(new_line)

		elif self.delete_check == 0 and self.add_check == 2:

			sql_request = "INSERT INTO f70 (protocol_id, date_of_issue, second_name, first_name, \
								surname, date_of_birth, address, \
								disability, CHI, outpatient_card_id, site, diagnosis, \
								code, doctor_second_name, type_f70) VALUES \
								(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

			new_line = []
			for i in range(1, 16):
				if i == self.id_wind_changing + 1: new_line += [self.new_data_for_changing]
				else: new_line += [self.all_data_saved[index_of_line][i]]

			if self.cheking_correctness_of_data(new_line) == False: return

			cursor.execute(sql_request, new_line)
			cursor.execute("SELECT COUNT(id) FROM f70")

			new_line = [cursor.fetchall()[0][0]] + new_line

			self.all_data_saved += [tuple(new_line)]
			self.page_label.setText("/ " + str(len(self.all_data_saved) // (self.count_of_lines_in_body+1) + 1))

		self.all_data_saved.sort(key = sort_func)
		self.dialog.hide()
		self.delete_body_data()
		self.add_data_from_db_in_body()

	def get_protocol_id_data(self, data):
		self.protocol_id_data = data

	def get_date_of_issue_data(self, data):
		try: self.date_of_issue_data = datetime.strptime(data, '%d.%m.%Y').date()
		except ValueError: self.date_of_issue_data = ''

	def get_second_name_data(self):
		if self.menu_widgets[2].toPlainText() != '': 
			self.second_name_data = self.menu_widgets[2].toPlainText().title() + "%"
		else: self.second_name_data = self.menu_widgets[2].toPlainText()

	def get_first_name_data(self):
		if self.menu_widgets[3].toPlainText() != '': self.first_name_data = self.menu_widgets[3].toPlainText().title() + "%"
		else: self.first_name_data = self.menu_widgets[3].toPlainText()

	def get_surname_data(self):
		if self.menu_widgets[4].toPlainText() != '': self.surname_data = self.menu_widgets[4].toPlainText().title() + "%"
		else: self.surname_data = self.menu_widgets[4].toPlainText()

	def get_date_of_birth_data(self, data):
		try: self.date_of_birth_data = datetime.strptime(data, '%d.%m.%Y').date()
		except ValueError: self.date_of_birth_data = ''

	def get_address_data(self):
		if self.menu_widgets[6].toPlainText() != '': 
			self.address_data = self.menu_widgets[6].toPlainText().capitalize() + "%"
		else: self.address_data = self.menu_widgets[6].toPlainText()

	def get_disability_data(self, data):
		if data == 0: self.disability_data = ''
		else: self.disability_data = self.disability_items[data]

	def get_CHI_data(self, data):
		self.CHI_data = data

	def get_outpatient_card_id_data(self, data):
		self.outpatient_card_id_data = data

	def get_site_data(self, data):
		self.site_data = data

	def get_diagnosis_data(self):
		if self.menu_widgets[11].toPlainText() != '': self.diagnosis_data = self.menu_widgets[11].toPlainText().lower() + "%"
		else: self.diagnosis_data = self.menu_widgets[11].toPlainText()

	def get_code_data(self, data):
		self.code_data = data

	def get_doctor_second_name_data(self, data):
		if len(data) == 0: self.doctor_second_name_data = ''
		else: self.doctor_second_name_data = data.title() + "%"

	def get_type_f70_data(self, data):
		if data == 0: self.type_f70_data = ''
		else: self.type_f70_data = self.type_f70_items[data]


	def paintEvent(self, e):
		self.qp = QPainter()
		self.qp.begin(self)
		self.qp.setBrush(QColor(150, 150, 150))
		self.qp.drawRect(0, 20, self.size().width(), 200)

		self.qp.setBrush(QColor(200, 200, 200))
		self.qp.drawRect(0, 220, self.size().width(), self.size().height())

		if self.selection_line != -1:
			self.qp.setBrush(QColor(5, 64, 4))
			self.qp.drawRect(0, self.line_changing * 30 + 252, self.size().width(), 30)


		self.qp.setBrush(QColor(0, 0, 0))
		self.qp.drawLine(0, 250, self.size().width(), 250)
		self.qp.end()

	def delete_body_data(self):

		for id_line in range(self.free_line_in_body):
			for wind in self.all_wind_in_body[id_line]:
				wind.hide()

		self.free_line_in_body = 0

	def the_search_button_was_clicked(self):
		self.search_button.setChecked(False)

		self.delete_body_data()
		self.all_data_saved = []
		self.page_in_body = 1
		self.page_label.setText("/ " + str(len(self.all_data_saved) // (self.count_of_lines_in_body+1) + 1))
		self.page_line.setText(str(self.page_in_body))

		for index in range(15):
			self.menu_labels[index].show()
			self.menu_widgets[index].show()
			self.menu_widgets[index].setStyleSheet("border: 2px solid black;")

		if self.mode_menu == 'adding':
			self.enter_add_button.hide()
		if self.mode_menu == 'reporting':
			self.enter_report_button.hide()
			self.report_period_line.hide()
			self.report_year_line.hide()
			self.report_type_line.hide()
			self.report_type_f70_line.hide()
			self.clear_all_menu_button.show()
			self.show_all_data_in_db_button.show()
			self.report_in_excel_button.hide()

			for index in range(7):
				self.report_info_lines[index].hide()
				self.report_info_labels[index].hide()

		self.mode_menu = 'searching'

		self.enter_search_button.show()

	def the_report_button_was_clicked(self):
		self.report_button.setChecked(False)

		self.delete_body_data()
		self.all_data_saved = []
		self.page_in_body = 1
		self.page_label.setText("/ " + str(len(self.all_data_saved) // (self.count_of_lines_in_body+1) + 1))
		self.page_line.setText(str(self.page_in_body))

		for index in range(15):
			self.menu_widgets[index].hide()
			self.menu_labels[index].hide()

		if self.mode_menu == 'adding':
			self.enter_add_button.hide()

		elif self.mode_menu == 'searching':
			self.enter_search_button.hide()

		self.mode_menu = 'reporting'

		self.report_period_line.show()
		self.report_year_line.show()
		self.report_type_line.show()
		self.enter_report_button.show()
		self.report_type_f70_line.show()
		self.report_in_excel_button.show()

		for index in range(7):
			self.report_info_lines[index].show()
			self.report_info_labels[index].show()

		self.clear_all_menu_button.hide()
		self.show_all_data_in_db_button.hide()

	def the_show_all_data_in_db_button_was_clicked(self):

		self.delete_body_data()
		self.all_data_saved = []

		sql_request = "SELECT * FROM f70 ORDER BY second_name, first_name, surname, date_of_issue DESC"

		cursor.execute(sql_request)
		self.all_data_saved = cursor.fetchall()

		self.page_in_body = 1
		self.page_label.setText("/ " + str(len(self.all_data_saved) // (self.count_of_lines_in_body+1) + 1))
		self.page_line.setText(str(self.page_in_body))
		self.add_data_from_db_in_body()

	def add_one_line_in_body(self, id_data):

		if self.free_line_in_body < self.count_of_lines_in_body:

			for i in range(15):
				if i == 1 or i == 5: 
					self.all_wind_in_body[self.free_line_in_body][i].   \
					setText(self.all_data_saved[id_data][i+1].strftime("%d.%m.%Y"))
				else: self.all_wind_in_body[self.free_line_in_body][i].setText(self.all_data_saved[id_data][i+1])
				self.all_wind_in_body[self.free_line_in_body][i].setAlignment(Qt.AlignmentFlag.AlignCenter)
				self.all_wind_in_body[self.free_line_in_body][i].show()

			self.free_line_in_body += 1

	def add_data_from_db_in_body(self):

		for id_data_line in range(self.count_of_lines_in_body*(self.page_in_body - 1), \
									min(self.count_of_lines_in_body*(self.page_in_body), len(self.all_data_saved))):
			self.add_one_line_in_body(id_data_line)

	def the_enter_search_button_was_clicked(self):
		self.enter_search_button.setChecked(False)

		sql_request = "SELECT id, protocol_id, date_of_issue, second_name, first_name, \
								surname, date_of_birth, address, \
								disability, CHI, outpatient_card_id, site, diagnosis, \
								code, doctor_second_name, type_f70 FROM f70 "

		first_add = True
		data = []
		for i in range(15):
			if getattr(self, self.english_names_coloms[i] + '_data') != '': 
				if first_add:
					first_add = False
					sql_request += "WHERE " + self.english_names_coloms[i]
				else: sql_request += "and " +  self.english_names_coloms[i]

				if i not in [1, 5]: sql_request += " LIKE %s "
				else: sql_request += " = %s "
				data += [getattr(self, self.english_names_coloms[i] + '_data')]

		sql_request += "ORDER BY second_name, first_name, surname, date_of_issue DESC "

		if not first_add: 
			self.delete_body_data()
			self.all_data_saved = []
			self.page_in_body = 1
			cursor.execute(sql_request, data)
			self.all_data_saved = cursor.fetchall()
			self.page_label.setText("/ " + str(len(self.all_data_saved) // (self.count_of_lines_in_body+1) + 1))
			self.page_line.setText(str(self.page_in_body))
			self.add_data_from_db_in_body()

	def dialog_menu_for_enter_add_button(self):

		self.dialog = QDialog(self)
		self.dialog.setGeometry(QRect(self.screen_size[0]//2 - 115, self.screen_size[1]//2 - 50, 230, 100))
		self.dialog.setWindowTitle("Предупреждение")

		label_1 = QLabel("Такие данные уже существуют.", self.dialog)
		label_1.setAlignment(Qt.AlignmentFlag.AlignCenter)
		label_1.setFont(self.font_arial_size10)
		label_1.setGeometry(QRect(0, 15, 230, 20))

		label_2 = QLabel("Все равно добавить?", self.dialog)
		label_2.setAlignment(Qt.AlignmentFlag.AlignCenter)
		label_2.setFont(self.font_arial_size10)
		label_2.setGeometry(QRect(0, 40, 230, 20))

		yes_button = QPushButton(self.dialog)
		yes_button.setCheckable(True)
		yes_button.setGeometry(QRect(20, self.dialog.size().height() - 25, 80, 20))
		yes_button.setText("Да")
		yes_button.clicked.connect(lambda: self.dialog_menu_for_enter_add_button_answer_data(1))

		no_button = QPushButton(self.dialog)
		no_button.setCheckable(True)
		no_button.setGeometry(QRect(130, self.dialog.size().height() - 25, 80, 20))
		no_button.setText("Нет")
		no_button.clicked.connect(lambda: self.dialog_menu_for_enter_add_button_answer_data(0))

		self.dialog.exec()

	def dialog_menu_for_enter_add_button_answer_data(self, data):
		self.dialog_menu_for_enter_add_button_answer = data
		self.dialog.hide()

	def the_enter_add_button_was_clicked(self):

		sql_request = "SELECT * FROM f70 "
		
		first_add = True
		data = []
		data_for_search = []
		for i in range(15):
			if getattr(self, self.english_names_coloms[i] + '_data') != '': 
				if first_add:
					first_add = False
					sql_request += "WHERE " + self.english_names_coloms[i] + " = %s "
				else: sql_request += "and " +  self.english_names_coloms[i] + " = %s "
				if i in [2, 3, 4, 6, 11]:
					data += [getattr(self, self.english_names_coloms[i] + '_data')[0:-1]]
					data_for_search += [getattr(self, self.english_names_coloms[i] + '_data')[0:-1]]
				else: 
					data += [getattr(self, self.english_names_coloms[i] + '_data')]
					data_for_search += [getattr(self, self.english_names_coloms[i] + '_data')]
			else: data += ['']

		if self.cheking_correctness_of_data(data):

			cursor.execute(sql_request, data_for_search)
			if len(cursor.fetchall()) > 0:
				self.dialog_menu_for_enter_add_button_answer = 0
				self.dialog_menu_for_enter_add_button()

			if self.dialog_menu_for_enter_add_button_answer:
				sql_request = "INSERT INTO f70 (protocol_id, date_of_issue, second_name, first_name, \
									surname, date_of_birth, address, \
									disability, CHI, outpatient_card_id, site, diagnosis, \
									code, doctor_second_name, type_f70) VALUES \
									(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
				if not first_add: 
					cursor.execute(sql_request, data)
					cursor.execute("SELECT COUNT(id) FROM f70")
					data = [cursor.fetchall()[0][0]] + data

					self.all_data_saved += [data]
					self.all_data_saved.sort(key = sort_func)

					self.page_label.setText("/ " + str(len(self.all_data_saved) // (self.count_of_lines_in_body+1) + 1))
					self.page_line.setText(str(self.page_in_body))

					self.delete_body_data()
					self.add_data_from_db_in_body()

		self.dialog_menu_for_enter_add_button_answer = 1

	def the_clear_all_menu_button_was_clicked(self):

		for id_widget in range(15):
			if id_widget not in [7, 14]: self.menu_widgets[id_widget].setText('')
			else: self.menu_widgets[id_widget].setCurrentIndex(0)

		self.protocol_id_data = ''
		self.date_of_issue_data = ''
		self.first_name_data = ''
		self.second_name_data = ''
		self.surname_data = ''
		self.date_of_birth_data = ''
		self.address_data = ''
		self.disability_data = ''
		self.CHI_data = ''
		self.outpatient_card_id_data = ''
		self.site_data = ''
		self.diagnosis_data = ''
		self.code_data = ''
		self.doctor_second_name_data = ''
		self.type_f70_data = ''

	def the_enter_report_button_was_clicked(self):
		sql_request = "SELECT * FROM f70 WHERE date_part('year', date_of_issue) = %s "

		if self.report_type == 'Взрослые': sql_request += "and date_part('year', age(date_of_birth)) >= 18 "
		elif self.report_type == 'Дети': sql_request += "and date_part('year', age(date_of_birth)) < 18 "
		else: sql_request += "and disability IN ('инв I гр.', 'инв II гр.', 'инв III гр.', 'реб. инв.') "

		if self.report_period_type == 'period': sql_request += "and date_part('mons', date_of_issue) BETWEEN 1 and %s "
		else: sql_request += "and date_part('mons', date_of_issue) = %s "

		cursor.execute(sql_request, [self.report_year, self.report_period])

		self.all_data_saved = []

		new_data = cursor.fetchall()
		self.report_info = [len(new_data), 0, 0, 0, 0, 0, 0]

		for line in new_data:
			self.report_info[self.type_f70_items.index(line[15])] += 1
			if self.report_type_f70 != '':
				if line[15] == self.report_type_f70: self.all_data_saved += [line]
			else: self.all_data_saved += [line]

		for i in range(7):
			self.report_info_lines[i].setText(str(self.report_info[i]))

		self.page_in_body = 1

		self.page_label.setText("/ " + str(len(self.all_data_saved) // (self.count_of_lines_in_body+1) + 1))
		self.page_line.setText(str(self.page_in_body))

		self.delete_body_data()
		self.add_data_from_db_in_body()

	def the_report_in_excel_button_was_clicked(self):
		self.the_enter_report_button_was_clicked()
		if len(self.all_data_saved) != 0:
			wb = Workbook()
			ws = wb.active

			title = "Отчет_" + self.report_type + "_"
			if self.report_period_type == "mons": title += self.period_report_items[self.report_period-1]
			else:
				if self.report_period == 3: title += self.period_report_items[12]
				elif self.report_period == 6: title += self.period_report_items[13]
				elif self.report_period == 9: title += self.period_report_items[14]
				elif self.report_period == 12: title += self.period_report_items[15]
			title += "_" + self.report_year
			ws.title = "Отчет"

			for i in range(16):
				if i == 0: ws.column_dimensions[letter(i+1)].width = 20
				else: ws.column_dimensions[letter(i+1)].width = 15

			ws['H1'] = self.report_type
			ws['I1'] = title.split('_')[2]
			ws['J1'] = int(self.report_year)
			ws['A2'] = ''

			for i in range(7):
				if i == 0: 
					ws['A'+str(i+3)] = "Всего"
					ws['A'+str(i+3)].font = Font(bold=True)
					ws['B'+str(i+3)].font = Font(bold=True)
				else: ws['A'+str(i+3)] = self.type_f70_items[i]
				ws['B'+str(i+3)] = self.report_info[i]

			ws['A10'] = ''

			sorted_data = {'органы дыхания' : [], 'органы опор-дв. апп.' : [], 'органы ССС' : [], \
							'органы эндок. сист.' : [], 'органы нерв. сист.' : [], 'другое' : []}
			for i in range(6):
				for line in self.all_data_saved:
					if line[15] == self.type_f70_items[i+1]:
						sorted_data[self.type_f70_items[i+1]] += [list(line)]

			y_coord = 11
			for key in sorted_data.keys():
				cnt_of_line = 1
				ws['I'+str(y_coord)] = key
				ws['I'+str(y_coord)].font = Font(bold=True)
				ws.append(['№'] + self.russian_names_coloms)
				y_coord += 2
				for line in sorted_data[key]:
					ws.append([cnt_of_line] + line[1:])
					cnt_of_line += 1
					y_coord += 1
				ws['A'+str(y_coord)] = ''
				y_coord += 1

			for row in ws.rows:
				for cell in row:
					cell.alignment = Alignment(horizontal="center", vertical="center")

			wb.save(title + '.xlsx')
			way = QFileDialog.getOpenFileName(self, "Выбрать файл", ".", "Excel (" + title + ".xlsx)")
			startfile(way[0])


	def the_page_down_button_was_clicked(self):

		if self.page_in_body * self.count_of_lines_in_body < len(self.all_data_saved):
			self.delete_body_data()
			self.page_in_body += 1
			self.page_line.setText(str(self.page_in_body))
			self.add_data_from_db_in_body()

	def the_page_up_button_was_clicked(self):

		if self.page_in_body > 1:
			self.delete_body_data()
			self.page_in_body -= 1
			self.page_line.setText(str(self.page_in_body))
			self.add_data_from_db_in_body()

	def change_page(self, data):

		try: int(data)
		except ValueError: return

		if int(data) <= len(self.all_data_saved) // (self.count_of_lines_in_body + 1) + 1:
			self.page_in_body = int(data)
			self.delete_body_data()
			self.add_data_from_db_in_body()
		else: self.page_line.setText(str(self.page_in_body))

	def the_add_button_was_clicked(self):
		self.add_button.setChecked(False)

		self.delete_body_data()
		self.all_data_saved = []
		self.page_in_body = 1
		self.page_label.setText("/ " + str(len(self.all_data_saved) // (self.count_of_lines_in_body+1) + 1))
		self.page_line.setText(str(self.page_in_body))

		for index in range(15):
			self.menu_widgets[index].show()
			self.menu_labels[index].show()
			if index not in [6, 8, 9, 10, 13]:
				self.menu_widgets[index].setStyleSheet("background-color: #57cf77; border: 2px solid black;")

		if self.mode_menu == 'searching':
			self.enter_search_button.hide()
		elif self.mode_menu == 'reporting':
			self.enter_report_button.hide()
			self.report_period_line.hide()
			self.report_year_line.hide()
			self.report_type_line.hide()
			self.report_type_f70_line.hide()
			self.clear_all_menu_button.show()
			self.show_all_data_in_db_button.show()
			self.report_in_excel_button.hide()

			for index in range(7):
				self.report_info_lines[index].hide()
				self.report_info_labels[index].hide()

		self.mode_menu = 'adding'
		self.enter_add_button.show()
		#self.get_file_name()

	def get_file_name(self):
		return QFileDialog.getOpenFileName(self, "Выбрать файл", ".", "Excel (*.xml *.xls *.xlsx)")
		# пример возвращения ('C:/Python programs/RL.xls', 'Excel (*.xml *.xls *.xlsx)')
		# если ничего не выбрал то ('', '')


if __name__ == "__main__":
	app = QApplication([])

	conn = psycopg2.connect(dbname="postgres", user="postgres", password="9", host="127.0.0.1")
	cursor = conn.cursor()

	conn.autocommit = True
	cursor.execute("CREATE TABLE IF NOT EXISTS f70 (id SERIAL PRIMARY KEY NOT NULL, \
												protocol_id VARCHAR(10) NOT NULL, \
												date_of_issue DATE NOT NULL,\
												first_name VARCHAR(20) NOT NULL,\
												second_name VARCHAR(20) NOT NULL,\
												surname VARCHAR(20) NOT NULL,\
												date_of_birth DATE NOT NULL,\
												address VARCHAR(100),\
												disability VARCHAR(11) NOT NULL,\
												CHI VARCHAR(16),\
												outpatient_card_id VARCHAR(4),\
												site VARCHAR(3),\
												diagnosis VARCHAR(50) NOT NULL,\
												code VARCHAR(10) NOT NULL,\
												doctor_second_name VARCHAR(20),\
												type_f70 VARCHAR(30) NOT NULL) ")

	window = MainWindow()
	window.show()

	app.exec()

	cursor.close()
	conn.close()